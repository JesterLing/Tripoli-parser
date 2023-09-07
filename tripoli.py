import json
import requests
import os
import openpyxl
import time
import re
import click
from bs4 import BeautifulSoup

CONFIG_FILE = 'config.json'
PARSED_DATA_FILE = 'result.json'
RESULT_FILE = 'result.xlsx'

SCRIPT_PATH = os.path.dirname(os.path.realpath(__file__))

def load_config():
  if os.path.isfile('{0}/{1}'.format(SCRIPT_PATH, CONFIG_FILE)):
    with open('{0}/{1}'.format(SCRIPT_PATH, CONFIG_FILE), 'r', encoding='utf-8') as f:
      return json.load(f)
  else:
    print('Не найден', CONFIG_FILE)
    exit()

def save_config(config):
  with open('{0}/{1}'.format(SCRIPT_PATH, CONFIG_FILE), 'w', encoding='utf-8') as f:
    json.dump(config, f, indent=2, ensure_ascii=False)

def check_auth(html_page):
  html = BeautifulSoup(html_page, 'html.parser')
  user_meta_tag = html.find('meta', attrs={'ng-init': re.compile(r'current_user_id')})
  user_id = int(user_meta_tag['ng-init'].split('=')[1])
  if user_id == -1: 
    print('Похоже что вы не авторизованы. Проверте cookies в', CONFIG_FILE)
    exit()

def download_pages(config):
  try:
    if not os.path.exists('{0}/pages'.format(SCRIPT_PATH)): os.mkdir('{0}/pages'.format(SCRIPT_PATH))
    print('Шаг 1. Скачиваем страницы с фермерскими хозяйствами')
    if config.get('current_page'):
      current_page = config['current_page']
    else:
      current_page = 1
    r = requests.get('https://tripoli.land/farmers/poltavskaya', params={'page': 1}, headers=config['headers'], cookies=config['cookies'])
    check_auth(r.text)
    page = BeautifulSoup(r.text, 'html.parser')
    pagination = page.css.select('.farmers-index-content .maincontent .pagination .col-sm-6.text-center a')
    last_page = int(pagination[-1].get_text())
    while current_page <= last_page:
      if os.path.isfile('{0}/pages/{1}.html'.format(SCRIPT_PATH, current_page)):
        print('Страница', current_page, 'существует, пропуск')
        current_page += 1
        continue
      r = requests.get('https://tripoli.land/farmers/poltavskaya', params={'page': current_page}, headers=config['headers'], cookies=config['cookies'])
      time.sleep(config['sleep'])
      if not r:
        print('Страница', current_page, 'ошибка, код', r.status_code)
      with open('{0}/pages/{1}.html'.format(SCRIPT_PATH, current_page), 'w', encoding='utf-8') as f:
        f.write(r.text)
      print('Страница [{0}/{1}] сохранена, {2} bytes'.format(current_page, last_page, len(r.content)))
      config['current_page'] = current_page
      current_page += 1
    config['current_job'] = 'step_2'
    config['current_page'] = 1
  except KeyboardInterrupt:
    print('Сохраниние состояния в', CONFIG_FILE, 'перед выходом...')
    save_config(config)

def parse_pages(config):
  r = requests.get('https://tripoli.land/farmers/poltavskaya', params={'page': 1}, headers=config['headers'], cookies=config['cookies'])
  check_auth(r.text)
  try:
    limit = False
    print('Шаг 2. Парсинг фермерских хозяйств на скачанных страницах и получение номеров')
    pages = [f for f in os.listdir('{0}/pages'.format(SCRIPT_PATH)) if os.path.isfile(os.path.join('{0}/pages'.format(SCRIPT_PATH), f))]
    if not pages:
      print('Отсутствуют скачаные страницы. Вернитесь к шагу 1')
      exit()
    pages.sort(key=lambda f: int(''.join(filter(str.isdigit, f))))
    if os.path.isfile('{0}/{1}'.format(SCRIPT_PATH, PARSED_DATA_FILE)):
      with open('{0}/{1}'.format(SCRIPT_PATH, PARSED_DATA_FILE), 'r', encoding='utf-8') as f:
        result = json.load(f)    
      print('Загрузка', PARSED_DATA_FILE)
    else:
      result = []
      print('Создание', PARSED_DATA_FILE)
    find = False if config.get('current_firm') else True
    for page in pages:
      if limit: break
      if config.get('current_page'):
        if int(os.path.splitext(page)[0]) < config['current_page']:
          continue
      config['current_page'] = int(os.path.splitext(page)[0])
      f = open('{0}/pages/{1}'.format(SCRIPT_PATH, page), 'r', encoding='utf-8')
      page_html = BeautifulSoup(f, 'html.parser')
      table = page_html.css.select('.farmers-index-content .maincontent .tripoli tbody')
      rows = table[0].find_all('tr')
      print('Парсинг страницы {:s}. Найдено'.format(page), len(rows), 'фирм')

      for idx, row in enumerate(rows):
        if limit: break
        cols = row.find_all('td')
        cleared = cols[0].find_all('div', class_='content-b')[1]
        name = cleared.find('span', class_='call-popup').get_text()
        new_badge = cleared.find('span', class_='interested')
        new_badge = True if new_badge != None else False
        director = cols[1].find('p').get_text()
        uniq_id = int(re.findall(r'[0-9]+', cols[1].get('ng-click'))[0])
        contacts = []
        if not find:
          if config['current_firm'] == uniq_id:
            find = True
          else:
            continue
        config['current_firm'] = uniq_id
        for index, firm in enumerate(result):
          if firm['id'] == uniq_id:
            result.pop(index)
            print('Фирма {0} с ID {1} уже существует. Перезаписываем'.format(firm['name'], firm['id']))
        
        print('Получение имен...', end='\r')
        r = requests.get('https://tripoli.land/profile/org_corrections', params={'org_id': uniq_id}, headers=config['headers'], cookies=config['cookies'])
        time.sleep(config['sleep'])
        names = r.json()
        print('Получение телефонов...', end='\r')
        r = requests.get('https://tripoli.land/farmers/{:d}/org_contacts'.format(uniq_id), headers=config['headers'], cookies=config['cookies'])
        time.sleep(config['sleep'])
        phones = r.json()
        for key, value in phones.items():
          if key == 'exit_u' or key == 'fax' or value == '' or value is None:
            continue
          if value == 'Превышен лимит':
            limit = True
          first_name = names[key]['first_name']
          last_name = names[key]['last_name']
          surname_name = names[key]['surname_name']
          fio = ' '.join(filter(None, (last_name, first_name, surname_name)))
          contacts.append([value, names[key]['position'], fio])
          
        r = requests.get('https://tripoli.land/profile/auxiliary_contacts', params={'catalog_org_id': uniq_id}, headers=config['headers'], cookies=config['cookies'])
        time.sleep(config['sleep'])
        auxiliarys = r.json()
        for auxiliary in auxiliarys:
          if auxiliary['value'] == '' or auxiliary['value'] is None:
            continue
          if auxiliary['value'] == 'Превышен лимит':
            limit = True
          first_name = auxiliary['first_name']
          last_name = auxiliary['last_name']
          surname_name = auxiliary['surname_name']
          fio = ' '.join(filter(None, (last_name, first_name, surname_name)))
          contacts.append([auxiliary['value'], auxiliary['position'], fio])

        result.append({'id': uniq_id, 'name': name, 'new': new_badge, 'director': director, 'contacts': contacts})
        print('Добавление фирмы {0} с ID {1}, получено {2} контактов. [{3}/{4}]'.format(name, uniq_id, len(contacts), idx + 1, len(rows)))
      f.close()
    if limit:
      print('Сайт начал возвращать "Превышен лимит". Подождите ~24 часа и запустите заново. Парсинг продолжиться с того же места')
      if click.confirm('Хотите создать Excel файл с уже существующими данными?', default=True): config['current_job'] = 'step_3'
    else:
      config['current_job'] = 'step_3'
      del config['current_page']
      del config['current_firm']
    with open('{0}/{1}'.format(SCRIPT_PATH, PARSED_DATA_FILE), 'w', encoding='utf-8') as f:
      json.dump(result, f, ensure_ascii=False)
    save_config(config)
  except KeyboardInterrupt:
    print('Сохраниние состояния в', CONFIG_FILE, 'и спарсеного в', PARSED_DATA_FILE, 'перед выходом...')
    with open('{0}/{1}'.format(SCRIPT_PATH, PARSED_DATA_FILE), 'w', encoding='utf-8') as f:
      json.dump(result, f, ensure_ascii=False)
    save_config(config)

def create_excel(config):
  print('Шаг 3. Создание Excel таблицы из данных')
  if config.get('current_page') or config.get('current_firm'):
    config['current_job'] = 'step_2'
    save_config(config)
  if not os.path.isfile('{0}/{1}'.format(os.path.dirname(os.path.realpath(__file__)), PARSED_DATA_FILE)):
    print('Спарсеные данные', PARSED_DATA_FILE, 'отсутствуют')
    exit()
  with open('{0}/{1}'.format(SCRIPT_PATH, PARSED_DATA_FILE), 'r', encoding='utf-8') as f:
    data = json.load(f)
  
  wb = openpyxl.Workbook()
  ws = wb.active
  fillTitle = openpyxl.styles.PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
  ws['A1'].value = 'Назва'
  ws['A1'].font = openpyxl.styles.Font(bold=True)
  ws['A1'].fill = fillTitle
  ws['A1'].alignment = openpyxl.styles.Alignment(vertical='center')
  ws.merge_cells('A1:A2')
  ws['C1'].value = 'Директор'
  ws['C1'].font = openpyxl.styles.Font(bold=True)
  ws['C1'].fill = fillTitle
  ws['C1'].alignment = openpyxl.styles.Alignment(vertical='center')
  ws.merge_cells('C1:C2')
  ws['D1'].value = 'Контакти'
  ws['D1'].font = openpyxl.styles.Font(bold=True)
  ws['D1'].fill = fillTitle
  ws.merge_cells('D1:F1')
  ws['D2'].value = 'Номер/Email'
  ws['D2'].font = openpyxl.styles.Font(bold=True)
  ws['D2'].fill = fillTitle
  ws['E2'].value = 'Позиція'
  ws['E2'].font = openpyxl.styles.Font(bold=True)
  ws['E2'].fill = fillTitle
  ws['F2'].value = 'П.І.Б.'
  ws['F2'].font = openpyxl.styles.Font(bold=True)
  ws['F2'].fill = fillTitle
  ws.column_dimensions['A'].width = 55
  ws.column_dimensions['B'].width = 5
  ws.column_dimensions['C'].width = 35
  ws.column_dimensions['D'].width = 25
  ws.column_dimensions['E'].width = 15
  ws.column_dimensions['F'].width = 25
  ws.merge_cells('B1:B2')
  ws['B1'].fill = fillTitle

  current_row = 3
  current_col = 1

  for idx, firm in enumerate(data):
    merge_row_start=current_row
    ws.cell(current_row, current_col).value = firm['name']
    if firm['new']:
      ws.cell(current_row, current_col+1).value = 'new'
    ws.cell(current_row, current_col+2).value = firm['director']
    for contact in firm['contacts']:
      ws.cell(current_row, current_col+3).value = contact[0]
      ws.cell(current_row, current_col+4).value = contact[1]
      ws.cell(current_row, current_col+5).value = contact[2]
      current_row += 1
    ws.cell(merge_row_start, 1).alignment = openpyxl.styles.Alignment(vertical='top')
    ws.cell(merge_row_start, 2).alignment = openpyxl.styles.Alignment(vertical='top')
    ws.cell(merge_row_start, 3).alignment = openpyxl.styles.Alignment(vertical='top')
    ws.merge_cells(start_row=merge_row_start, start_column=1, end_row=current_row, end_column=1)
    ws.merge_cells(start_row=merge_row_start, start_column=2, end_row=current_row, end_column=2)
    ws.merge_cells(start_row=merge_row_start, start_column=3, end_row=current_row, end_column=3)
    current_row += 1
    print('Генерация {:.0%}'.format(idx/len(data)), end='\r')

  wb.save('{0}/{1}'.format(SCRIPT_PATH, RESULT_FILE))
  print('Завершено!', RESULT_FILE, 'в папке со скриптом')

def main():
  config = load_config()
  if not config.get('headers'):
    print('Нет заголовков в', CONFIG_FILE)
    exit()
  if not config.get('cookies'):
    print('Нет cookies в', CONFIG_FILE)
    exit()
  if not config.get('current_job'): config['current_job'] = 'step_1'
  if config['current_job'] not in ('step_1', 'step_2', 'step_3'):
    print('Значение current_job в', CONFIG_FILE, 'недопустимое')
    exit()

  if config.get('current_job') == 'step_1':
    download_pages(config)
  if config.get('current_job') == 'step_2':
    parse_pages(config)
  if config.get('current_job') == 'step_3':
    create_excel(config)

  save_config(config)

if __name__ == '__main__':
  main()